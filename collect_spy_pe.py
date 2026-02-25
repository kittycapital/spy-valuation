#!/usr/bin/env python3
"""
SPY Valuation Dashboard - Data Collector
Fetches Shiller CAPE/P/E data + SPY prices and builds JSON for the dashboard.

Usage:
  python collect_spy_pe.py --full    # Initial: Shiller Excel + SPY CSV
  python collect_spy_pe.py --update  # Daily: Yahoo Finance SPY price only
  python collect_spy_pe.py --repair  # Re-download Shiller + rebuild
"""

import json
import csv
import sys
import os
import time
import struct
from pathlib import Path
from datetime import datetime, timezone
from urllib.request import urlopen, Request
from urllib.error import URLError

DATA_DIR = Path("data")

SHILLER_URL = "http://www.econ.yale.edu/~shiller/data/ie_data.xls"

# Valuation bands for CAPE
CAPE_BANDS = [
    {"label": "매우 저평가", "min": 0, "max": 15, "color": "rgba(34,197,94,0.08)", "border": "rgba(34,197,94,0.3)"},
    {"label": "저평가", "min": 15, "max": 20, "color": "rgba(59,130,246,0.06)", "border": "rgba(59,130,246,0.2)"},
    {"label": "적정", "min": 20, "max": 25, "color": "rgba(107,114,128,0.04)", "border": "rgba(107,114,128,0.15)"},
    {"label": "고평가", "min": 25, "max": 30, "color": "rgba(245,158,11,0.06)", "border": "rgba(245,158,11,0.2)"},
    {"label": "과열", "min": 30, "max": 999, "color": "rgba(239,68,68,0.06)", "border": "rgba(239,68,68,0.2)"},
]

# Valuation bands for Trailing P/E
PE_BANDS = [
    {"label": "매우 저평가", "min": 0, "max": 12, "color": "rgba(34,197,94,0.08)", "border": "rgba(34,197,94,0.3)"},
    {"label": "저평가", "min": 12, "max": 15, "color": "rgba(59,130,246,0.06)", "border": "rgba(59,130,246,0.2)"},
    {"label": "적정", "min": 15, "max": 20, "color": "rgba(107,114,128,0.04)", "border": "rgba(107,114,128,0.15)"},
    {"label": "고평가", "min": 20, "max": 25, "color": "rgba(245,158,11,0.06)", "border": "rgba(245,158,11,0.2)"},
    {"label": "과열", "min": 25, "max": 999, "color": "rgba(239,68,68,0.06)", "border": "rgba(239,68,68,0.2)"},
]


def download_file(url, dest):
    """Download a file from URL."""
    print(f"  Downloading {url}...")
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urlopen(req, timeout=60)
        with open(dest, "wb") as f:
            f.write(resp.read())
        print(f"  Saved to {dest} ({dest.stat().st_size / 1024:.1f} KB)")
        return True
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


def parse_xls_basic(filepath):
    """
    Parse old-format .xls (BIFF8) to extract Shiller data.
    Tries xlrd first, then libreoffice+openpyxl fallback.
    """
    # Try xlrd (works in GitHub Actions)
    try:
        import xlrd
        return parse_xls_xlrd(filepath)
    except ImportError:
        pass

    # Try libreoffice conversion (works locally)
    import subprocess
    xlsx_path = filepath.with_suffix(".xlsx")
    try:
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx",
             "--outdir", str(filepath.parent), str(filepath)],
            capture_output=True, timeout=60
        )
        if xlsx_path.exists():
            return parse_xlsx(xlsx_path)
    except Exception as e:
        print(f"  LibreOffice conversion failed: {e}")

    # Fallback: existing CSV
    csv_fallback = DATA_DIR / "shiller_data.csv"
    if csv_fallback.exists():
        print("  Using existing shiller_data.csv")
        return parse_shiller_csv(csv_fallback)

    return None


def parse_xls_xlrd(filepath):
    """Parse xls file using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_name("Data")

    data = []
    for r in range(8, ws.nrows):
        date_val = ws.cell_value(r, 0)
        price = ws.cell_value(r, 1)
        earnings = ws.cell_value(r, 3)
        cape = ws.cell_value(r, 12)

        if not date_val or not price:
            continue
        if not isinstance(date_val, (int, float)):
            continue

        year = int(date_val)
        month = round((date_val - year) * 100)
        if month == 0:
            month = 1
        date_str = f"{year}-{month:02d}"

        trailing_pe = None
        if isinstance(earnings, (int, float)) and earnings > 0:
            trailing_pe = round(price / earnings, 2)

        cape_val = None
        if isinstance(cape, (int, float)):
            cape_val = round(cape, 2)

        data.append({
            "date": date_str,
            "sp500": round(price, 2),
            "earnings": round(earnings, 4) if isinstance(earnings, (int, float)) else None,
            "cape": cape_val,
            "trailing_pe": trailing_pe,
        })

    print(f"  Parsed {len(data)} monthly records from Shiller data (xlrd)")
    return data


def parse_xlsx(filepath):
    """Parse xlsx file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not available, trying alternative...")
        return None

    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["Data"]

    # Find header row (row 8 has column names)
    # Date, P, D, E, CPI, Fraction, Rate GS10, Real Price, Real Dividend,
    # Real Return Price, Real Earnings, Scaled Earnings, CAPE, _, TR CAPE
    data = []
    for r in range(9, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        price = ws.cell(r, 2).value
        earnings = ws.cell(r, 4).value
        cape = ws.cell(r, 13).value

        if date_val is None or price is None:
            continue

        # Convert date like 2023.09 to "2023-09"
        year = int(date_val)
        month = round((date_val - year) * 100)
        if month == 0:
            month = 1
        date_str = f"{year}-{month:02d}"

        # Calculate trailing P/E
        trailing_pe = None
        if earnings and earnings > 0:
            trailing_pe = round(price / earnings, 2)

        # CAPE
        cape_val = None
        if cape and cape != "NA" and isinstance(cape, (int, float)):
            cape_val = round(cape, 2)

        data.append({
            "date": date_str,
            "sp500": round(price, 2),
            "earnings": round(earnings, 4) if earnings else None,
            "cape": cape_val,
            "trailing_pe": trailing_pe,
        })

    print(f"  Parsed {len(data)} monthly records from Shiller data")
    return data


def parse_shiller_csv(filepath):
    """Parse pre-converted CSV of Shiller data."""
    data = []
    with open(filepath, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                data.append({
                    "date": row["date"],
                    "sp500": float(row["sp500"]) if row.get("sp500") else None,
                    "earnings": float(row["earnings"]) if row.get("earnings") else None,
                    "cape": float(row["cape"]) if row.get("cape") else None,
                    "trailing_pe": float(row["trailing_pe"]) if row.get("trailing_pe") else None,
                })
            except (ValueError, KeyError):
                continue
    print(f"  Parsed {len(data)} records from CSV")
    return data


def fetch_multpl_data():
    """
    Scrape recent CAPE and trailing P/E from multpl.com
    to supplement Shiller data which may lag by months.
    Uses multiple parsing strategies for robustness.
    """
    import re

    cape_data = {}
    pe_data = {}

    urls = [
        ("https://www.multpl.com/shiller-pe/table/by-month", "cape"),
        ("https://www.multpl.com/s-p-500-pe-ratio/table/by-month", "pe"),
    ]

    months_map = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
    }

    # Reasonable value ranges for validation
    VALID_RANGE = {"cape": (5, 200), "pe": (3, 200)}

    for url, metric in urls:
        print(f"  Fetching {metric} from multpl.com...")
        try:
            req = Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            })
            resp = urlopen(req, timeout=30)
            html = resp.read().decode("utf-8", errors="ignore")

            target = cape_data if metric == "cape" else pe_data
            vmin, vmax = VALID_RANGE[metric]
            skipped = 0

            # Strategy 1: find each <tr> and extract the two <td> cells
            tr_blocks = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE)

            for tr in tr_blocks:
                # Extract all <td> contents
                tds = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.DOTALL | re.IGNORECASE)
                if len(tds) < 2:
                    continue

                # First td: date like "Jan 1, 2025" or with links
                date_text = re.sub(r'<[^>]+>', '', tds[0]).strip()
                # Second td: value like "38.25" or "†38.25" or with HTML entities
                val_text = re.sub(r'<[^>]+>', '', tds[1]).strip()
                # Also clean HTML entities
                val_text = val_text.replace('&dagger;', '').replace('†', '').replace('\xa0', '').strip()

                # Parse date - try multiple formats
                date_match = re.match(r'([A-Z][a-z]{2})\s+\d{1,2},?\s*(\d{4})', date_text)
                if not date_match:
                    # Try format: "2025-01-01" or "1, Jan 2025"
                    date_match2 = re.search(r'(\d{4})-(\d{2})', date_text)
                    if date_match2:
                        year = date_match2.group(1)
                        mon = date_match2.group(2)
                    else:
                        continue
                else:
                    mon = months_map.get(date_match.group(1))
                    year = date_match.group(2)

                if not mon or not year:
                    continue

                # Parse value: strip non-numeric prefix (†, &dagger;, spaces, etc.)
                val_match = re.search(r'(\d+\.?\d*)', val_text)
                if not val_match:
                    continue

                value = float(val_match.group(1))

                # Validation: reject values outside reasonable range
                if value < vmin or value > vmax:
                    skipped += 1
                    continue

                target[f"{year}-{mon}"] = value

            # Strategy 2: if we got very few results, try alternative parsing
            if len(target) < 12:
                print(f"    Strategy 1 got only {len(target)} results, trying alternative parsing...")
                # Try finding the table by id
                table_match = re.search(r'<table[^>]*id=["\']datatable["\'][^>]*>(.*?)</table>', html, re.DOTALL | re.IGNORECASE)
                if table_match:
                    table_html = table_match.group(1)
                    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table_html, re.DOTALL | re.IGNORECASE)
                    for tr in rows:
                        tds = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.DOTALL | re.IGNORECASE)
                        if len(tds) >= 2:
                            date_text = re.sub(r'<[^>]+>', '', tds[0]).strip()
                            val_text = re.sub(r'<[^>]+>', '', tds[1]).strip().replace('&dagger;', '').replace('†', '').replace('\xa0', '')
                            date_match = re.match(r'([A-Z][a-z]{2})\s+\d{1,2},?\s*(\d{4})', date_text)
                            if date_match:
                                mon = months_map.get(date_match.group(1))
                                year = date_match.group(2)
                                val_match = re.search(r'(\d+\.?\d*)', val_text)
                                if mon and year and val_match:
                                    value = float(val_match.group(1))
                                    if vmin <= value <= vmax:
                                        target[f"{year}-{mon}"] = value

            if skipped:
                print(f"    Skipped {skipped} invalid values for {metric}")
            print(f"    Got {len(target)} months for {metric}")
        except Exception as e:
            print(f"    Failed: {e}")

    return cape_data, pe_data


def merge_multpl_into_shiller(shiller_data, cape_data, pe_data):
    """Merge multpl.com data into shiller data, filling gaps."""
    existing_dates = {e["date"] for e in shiller_data}
    all_new_dates = sorted((set(cape_data.keys()) | set(pe_data.keys())) - existing_dates)

    added = 0
    for d in all_new_dates:
        if d < "1900":
            continue
        cape_val = round(cape_data[d], 2) if d in cape_data else None
        pe_val = round(pe_data[d], 2) if d in pe_data else None
        # Extra validation
        if cape_val and cape_val > 200:
            cape_val = None
        if pe_val and pe_val > 200:
            pe_val = None
        shiller_data.append({
            "date": d, "sp500": None, "earnings": None,
            "cape": cape_val,
            "trailing_pe": pe_val,
        })
        added += 1

    # Fill missing values in existing entries
    for entry in shiller_data:
        d = entry["date"]
        if entry["cape"] is None and d in cape_data:
            val = round(cape_data[d], 2)
            if val <= 200:
                entry["cape"] = val
        if entry["trailing_pe"] is None and d in pe_data:
            val = round(pe_data[d], 2)
            if val <= 200:
                entry["trailing_pe"] = val
        # Clean any existing bad values
        if entry.get("cape") and entry["cape"] > 200:
            entry["cape"] = None
        if entry.get("trailing_pe") and entry["trailing_pe"] > 200:
            entry["trailing_pe"] = None

    shiller_data.sort(key=lambda x: x["date"])
    if added > 0:
        print(f"  Added {added} new months from multpl.com")
    return shiller_data


def fill_shiller_gaps(shiller_data, spy_daily):
    """
    Fill gaps in shiller monthly data with placeholder entries.
    Uses SPY daily prices to fill sp500 field for missing months.
    This ensures continuous timeline for charts even when CAPE/PE data is missing.
    """
    existing_dates = {e["date"] for e in shiller_data}

    # Find date range
    all_dates = sorted(existing_dates)
    if not all_dates:
        return shiller_data

    start_y, start_m = int(all_dates[0][:4]), int(all_dates[0][5:7])
    end_y, end_m = int(all_dates[-1][:4]), int(all_dates[-1][5:7])

    # Also check spy_daily for the latest date
    if spy_daily:
        latest_spy = sorted(spy_daily.keys())[-1]
        spy_y, spy_m = int(latest_spy[:4]), int(latest_spy[5:7])
        if spy_y * 12 + spy_m > end_y * 12 + end_m:
            end_y, end_m = spy_y, spy_m

    added = 0
    y, m = start_y, start_m
    while y * 12 + m <= end_y * 12 + end_m:
        date_str = f"{y}-{m:02d}"
        if date_str not in existing_dates and date_str >= "1900-01":
            # Find SPY price for this month (use last trading day)
            sp500_price = None
            month_days = sorted([k for k in spy_daily if k.startswith(date_str)])
            if month_days:
                sp500_price = spy_daily[month_days[-1]]

            shiller_data.append({
                "date": date_str,
                "sp500": round(sp500_price, 2) if sp500_price else None,
                "earnings": None,
                "cape": None,
                "trailing_pe": None,
            })
            added += 1

        m += 1
        if m > 12:
            m = 1
            y += 1

    if added > 0:
        shiller_data.sort(key=lambda x: x["date"])
        print(f"  Filled {added} gap months with placeholder entries")

    return shiller_data


def load_spy_csv():
    """Load SPY daily prices from CSV."""
    csv_path = DATA_DIR / "SPY.csv"
    if not csv_path.exists():
        csv_path = Path("SPY.csv")
    if not csv_path.exists():
        print("  SPY.csv not found")
        return {}

    prices = {}
    with open(csv_path, "r") as f:
        header = True
        for line in f:
            if header:
                header = False
                continue
            parts = line.strip().split(",")
            if len(parts) >= 2:
                try:
                    date_str = parts[0].strip()
                    close = float(parts[1])
                    prices[date_str] = round(close, 2)
                except (ValueError, IndexError):
                    continue
    print(f"  Loaded {len(prices)} daily SPY prices from CSV")
    return prices


def fetch_spy_today():
    """Fetch latest SPY prices from Yahoo Finance (multiple API endpoints)."""
    prices = {}

    # Method 1: Yahoo v8 chart API
    print("  Fetching latest SPY price from Yahoo Finance...")
    try:
        url = "https://query1.finance.yahoo.com/v8/finance/chart/SPY?range=5d&interval=1d"
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urlopen(req, timeout=15)
        data = json.loads(resp.read())

        result = data["chart"]["result"][0]
        timestamps = result["timestamp"]
        closes = result["indicators"]["quote"][0]["close"]

        for ts, close in zip(timestamps, closes):
            if close is not None:
                date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
                prices[date_str] = round(close, 2)

        print(f"  Got {len(prices)} recent SPY prices (v8 API)")
    except Exception as e:
        print(f"  Yahoo v8 failed: {e}")

    # Method 2: Yahoo v7 download API (fallback)
    if not prices:
        try:
            print("  Trying Yahoo v7 download API...")
            end_ts = int(time.time())
            start_ts = end_ts - 7 * 86400
            url = f"https://query1.finance.yahoo.com/v7/finance/download/SPY?period1={start_ts}&period2={end_ts}&interval=1d&events=history"
            req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
            resp = urlopen(req, timeout=15)
            csv_text = resp.read().decode("utf-8")
            for line in csv_text.strip().split("\n")[1:]:
                parts = line.split(",")
                if len(parts) >= 5 and parts[4] != "null":
                    try:
                        prices[parts[0]] = round(float(parts[4]), 2)
                    except ValueError:
                        continue
            print(f"  Got {len(prices)} recent SPY prices (v7 download)")
        except Exception as e:
            print(f"  Yahoo v7 failed: {e}")

    # Method 3: Yahoo v10 quote API (single latest price)
    if not prices:
        try:
            print("  Trying Yahoo v10 quote API...")
            url = "https://query2.finance.yahoo.com/v10/finance/quoteSummary/SPY?modules=price"
            req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
            resp = urlopen(req, timeout=15)
            data = json.loads(resp.read())
            price_data = data["quoteSummary"]["result"][0]["price"]
            close = price_data.get("regularMarketPrice", {}).get("raw")
            mkt_time = price_data.get("regularMarketTime")
            if close and mkt_time:
                date_str = datetime.fromtimestamp(mkt_time, tz=timezone.utc).strftime("%Y-%m-%d")
                prices[date_str] = round(close, 2)
                print(f"  Got latest SPY price: ${close:.2f} on {date_str}")
        except Exception as e:
            print(f"  Yahoo v10 failed: {e}")

    if not prices:
        print("  WARNING: All Yahoo Finance APIs failed. SPY prices not updated.")

    return prices


def update_spy_csv(new_prices):
    """Append new dates to SPY.csv if not already present."""
    csv_path = DATA_DIR / "SPY.csv"
    if not csv_path.exists():
        print("  SPY.csv not found, skipping CSV update")
        return

    # Read existing dates
    existing_dates = set()
    with open(csv_path, "r") as f:
        for i, line in enumerate(f):
            if i == 0:
                continue
            parts = line.strip().split(",")
            if parts:
                existing_dates.add(parts[0])

    # Append new dates
    added = 0
    with open(csv_path, "a") as f:
        for date_str in sorted(new_prices.keys()):
            if date_str not in existing_dates:
                price = new_prices[date_str]
                f.write(f"{date_str},{price},{price},{price},{price},0\n")
                added += 1

    if added:
        print(f"  Appended {added} new dates to SPY.csv")
    else:
        print(f"  SPY.csv already up to date")


def calculate_percentile(value, all_values):
    """Calculate historical percentile for a value."""
    if not all_values or value is None:
        return None
    below = sum(1 for v in all_values if v <= value)
    return round(below / len(all_values) * 100, 1)


def calculate_forward_returns(shiller_data, spy_daily):
    """
    Calculate average forward 10-year returns by CAPE range.
    Uses monthly S&P 500 data from Shiller.
    """
    # Build monthly price series
    monthly_prices = {}
    for entry in shiller_data:
        if entry["sp500"]:
            monthly_prices[entry["date"]] = entry["sp500"]

    sorted_dates = sorted(monthly_prices.keys())
    results = {}

    for i, date_str in enumerate(sorted_dates):
        cape = None
        for entry in shiller_data:
            if entry["date"] == date_str:
                cape = entry.get("cape")
                break

        if cape is None:
            continue

        # Find price 10 years later
        year = int(date_str[:4])
        month = int(date_str[5:7])
        future_date = f"{year + 10}-{month:02d}"

        if future_date in monthly_prices:
            start_price = monthly_prices[date_str]
            end_price = monthly_prices[future_date]
            annual_return = ((end_price / start_price) ** (1 / 10) - 1) * 100

            # Categorize by CAPE band
            band_key = None
            if cape < 15:
                band_key = "<15"
            elif cape < 20:
                band_key = "15-20"
            elif cape < 25:
                band_key = "20-25"
            elif cape < 30:
                band_key = "25-30"
            else:
                band_key = "30+"

            if band_key not in results:
                results[band_key] = []
            results[band_key].append(annual_return)

    # Calculate averages
    summary = {}
    for band, returns in results.items():
        summary[band] = {
            "avg": round(sum(returns) / len(returns), 1),
            "min": round(min(returns), 1),
            "max": round(max(returns), 1),
            "count": len(returns),
        }

    return summary


def build_chart_data(shiller_data, spy_daily):
    """Build the final JSON data for the dashboard."""

    # All CAPE values for percentile (filter out bad values)
    all_capes = [e["cape"] for e in shiller_data if e["cape"] is not None and e["cape"] <= 200]
    all_pes = [e["trailing_pe"] for e in shiller_data if e["trailing_pe"] is not None and e["trailing_pe"] <= 200]

    # Latest values (skip bad values > 200)
    latest_shiller = None
    for e in reversed(shiller_data):
        if e["cape"] is not None and e["cape"] <= 200:
            latest_shiller = e
            break

    latest_pe_entry = None
    for e in reversed(shiller_data):
        if e["trailing_pe"] is not None and e["trailing_pe"] <= 200:
            latest_pe_entry = e
            break

    # Latest SPY price
    spy_sorted = sorted(spy_daily.keys())
    latest_spy_price = spy_daily[spy_sorted[-1]] if spy_sorted else None
    latest_spy_date = spy_sorted[-1] if spy_sorted else None

    # Forward returns
    forward_returns = calculate_forward_returns(shiller_data, spy_daily)

    chart_data = {
        "shiller": [],
        "spy_daily": {},
        "summary": {
            "latest_cape": latest_shiller["cape"] if latest_shiller else None,
            "latest_cape_date": latest_shiller["date"] if latest_shiller else None,
            "cape_percentile": calculate_percentile(
                latest_shiller["cape"] if latest_shiller else None, all_capes
            ),
            "latest_pe": latest_pe_entry["trailing_pe"] if latest_pe_entry else None,
            "latest_pe_date": latest_pe_entry["date"] if latest_pe_entry else None,
            "pe_percentile": calculate_percentile(
                latest_pe_entry["trailing_pe"] if latest_pe_entry else None, all_pes
            ),
            "latest_spy": latest_spy_price,
            "latest_spy_date": latest_spy_date,
            "cape_median": round(sorted(all_capes)[len(all_capes) // 2], 1) if all_capes else None,
            "cape_mean": round(sum(all_capes) / len(all_capes), 1) if all_capes else None,
            "pe_median": round(sorted(all_pes)[len(all_pes) // 2], 1) if all_pes else None,
            "pe_mean": round(sum(all_pes) / len(all_pes), 1) if all_pes else None,
        },
        "forward_returns": forward_returns,
        "cape_bands": CAPE_BANDS,
        "pe_bands": PE_BANDS,
        "metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "shiller_records": len(shiller_data),
            "spy_daily_records": len(spy_daily),
        },
    }

    # Trim shiller data to 1900+ for chart (pre-1900 is less relevant)
    for entry in shiller_data:
        if entry["date"] >= "1900":
            # Clean bad values before adding to chart data
            clean = dict(entry)
            if clean.get("cape") and clean["cape"] > 200:
                clean["cape"] = None
            if clean.get("trailing_pe") and clean["trailing_pe"] > 200:
                clean["trailing_pe"] = None
            chart_data["shiller"].append(clean)

    # SPY daily (only keep monthly samples for reasonable JSON size)
    for date_str in sorted(spy_daily.keys()):
        chart_data["spy_daily"][date_str] = spy_daily[date_str]

    return chart_data


def save_shiller_csv(shiller_data):
    """Save parsed Shiller data as CSV for future fallback."""
    csv_path = DATA_DIR / "shiller_data.csv"
    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["date", "sp500", "earnings", "cape", "trailing_pe"])
        writer.writeheader()
        writer.writerows(shiller_data)
    print(f"  Saved Shiller CSV backup: {csv_path}")


def run_full():
    """Full data collection: Shiller Excel + SPY CSV."""
    DATA_DIR.mkdir(exist_ok=True)

    # 1. Download and parse Shiller data
    print("\n=== Downloading Shiller CAPE Data ===")
    xls_path = DATA_DIR / "ie_data.xls"
    shiller_data = None

    # Try download first
    if download_file(SHILLER_URL, xls_path):
        shiller_data = parse_xls_basic(xls_path)

    # Fallback to local file in data/ or root
    if not shiller_data and xls_path.exists():
        print("  Using local file: " + str(xls_path))
        shiller_data = parse_xls_basic(xls_path)

    if not shiller_data:
        local_xls = Path("ie_data.xls")
        if local_xls.exists():
            shiller_data = parse_xls_basic(local_xls)

    if not shiller_data:
        print("ERROR: Cannot get Shiller data")
        sys.exit(1)

    save_shiller_csv(shiller_data)

    # 1b. Supplement with multpl.com for recent months
    print("\n=== Fetching Recent P/E Data from multpl.com ===")
    cape_data, pe_data = fetch_multpl_data()
    if cape_data or pe_data:
        shiller_data = merge_multpl_into_shiller(shiller_data, cape_data, pe_data)

    # 2. Load SPY prices
    print("\n=== Loading SPY Prices ===")
    spy_daily = load_spy_csv()
    today_prices = fetch_spy_today()
    spy_daily.update(today_prices)

    # Update SPY.csv with new prices
    if today_prices:
        update_spy_csv(today_prices)

    # 2b. Fill gaps in shiller data with SPY prices
    print("\n=== Filling Data Gaps ===")
    shiller_data = fill_shiller_gaps(shiller_data, spy_daily)

    # 3. Build and save
    print("\n=== Building Chart Data ===")
    chart_data = build_chart_data(shiller_data, spy_daily)

    output_path = DATA_DIR / "spy_valuation.json"
    with open(output_path, "w") as f:
        json.dump(chart_data, f)
    print(f"\nData saved: {output_path} ({output_path.stat().st_size / 1024:.1f} KB)")


def run_update():
    """Daily update: add latest SPY price + refresh P/E from multpl.com."""
    data_path = DATA_DIR / "spy_valuation.json"
    if not data_path.exists():
        print("No existing data. Run --full first.")
        sys.exit(1)

    with open(data_path) as f:
        chart_data = json.load(f)

    # Fetch latest SPY prices
    today_prices = fetch_spy_today()
    if today_prices:
        chart_data["spy_daily"].update(today_prices)
        spy_sorted = sorted(chart_data["spy_daily"].keys())
        chart_data["summary"]["latest_spy"] = chart_data["spy_daily"][spy_sorted[-1]]
        chart_data["summary"]["latest_spy_date"] = spy_sorted[-1]
        # Also update SPY.csv
        update_spy_csv(today_prices)

    # Refresh P/E data from multpl.com
    cape_data, pe_data = fetch_multpl_data()
    if cape_data or pe_data:
        shiller = chart_data.get("shiller", [])
        existing_dates = {e["date"] for e in shiller}

        # Add new months
        spy_daily = chart_data.get("spy_daily", {})
        for d in sorted((set(cape_data.keys()) | set(pe_data.keys())) - existing_dates):
            if d < "1900":
                continue
            # Try to fill sp500 price from spy_daily (use last day of month)
            sp500_price = None
            month_prefix = d  # e.g. "2024-06"
            month_days = sorted([k for k in spy_daily if k.startswith(month_prefix)])
            if month_days:
                sp500_price = spy_daily[month_days[-1]]
            cape_val = round(cape_data[d], 2) if d in cape_data else None
            pe_val = round(pe_data[d], 2) if d in pe_data else None
            if cape_val and cape_val > 200: cape_val = None
            if pe_val and pe_val > 200: pe_val = None
            shiller.append({
                "date": d, "sp500": sp500_price, "earnings": None,
                "cape": cape_val, "trailing_pe": pe_val,
            })

        # Fill missing in existing + clean bad values
        for entry in shiller:
            d = entry["date"]
            if entry.get("cape") is None and d in cape_data:
                val = round(cape_data[d], 2)
                if val <= 200:
                    entry["cape"] = val
            if entry.get("trailing_pe") is None and d in pe_data:
                val = round(pe_data[d], 2)
                if val <= 200:
                    entry["trailing_pe"] = val
            # Clean existing bad values
            if entry.get("cape") and entry["cape"] > 200:
                entry["cape"] = None
            if entry.get("trailing_pe") and entry["trailing_pe"] > 200:
                entry["trailing_pe"] = None

        shiller.sort(key=lambda x: x["date"])
        chart_data["shiller"] = shiller

        # Update summary with latest values (only valid ones)
        all_capes = [e["cape"] for e in shiller if e.get("cape") is not None and e["cape"] <= 200]
        all_pes = [e["trailing_pe"] for e in shiller if e.get("trailing_pe") is not None and e["trailing_pe"] <= 200]

        for e in reversed(shiller):
            if e.get("cape") is not None and e["cape"] <= 200:
                chart_data["summary"]["latest_cape"] = e["cape"]
                chart_data["summary"]["latest_cape_date"] = e["date"]
                below = sum(1 for v in all_capes if v <= e["cape"])
                chart_data["summary"]["cape_percentile"] = round(below / len(all_capes) * 100, 1)
                break
        for e in reversed(shiller):
            if e.get("trailing_pe") is not None and e["trailing_pe"] <= 200:
                chart_data["summary"]["latest_pe"] = e["trailing_pe"]
                chart_data["summary"]["latest_pe_date"] = e["date"]
                below = sum(1 for v in all_pes if v <= e["trailing_pe"])
                chart_data["summary"]["pe_percentile"] = round(below / len(all_pes) * 100, 1)
                break

        # Recalculate median/mean with clean data
        if all_capes:
            chart_data["summary"]["cape_median"] = round(sorted(all_capes)[len(all_capes)//2], 1)
            chart_data["summary"]["cape_mean"] = round(sum(all_capes)/len(all_capes), 1)
        if all_pes:
            chart_data["summary"]["pe_median"] = round(sorted(all_pes)[len(all_pes)//2], 1)
            chart_data["summary"]["pe_mean"] = round(sum(all_pes)/len(all_pes), 1)

    chart_data["metadata"]["generated_at"] = datetime.now(timezone.utc).isoformat()
    chart_data["metadata"]["mode"] = "update"

    output_path = DATA_DIR / "spy_valuation.json"
    with open(output_path, "w") as f:
        json.dump(chart_data, f)
    print(f"Updated: {output_path}")


def run_repair():
    """Re-download Shiller data and rebuild everything."""
    run_full()


if __name__ == "__main__":
    if "--full" in sys.argv:
        run_full()
    elif "--update" in sys.argv:
        run_update()
    elif "--repair" in sys.argv:
        run_repair()
    else:
        print("Usage:")
        print("  python collect_spy_pe.py --full    # Initial full collection")
        print("  python collect_spy_pe.py --update  # Daily SPY price update")
        print("  python collect_spy_pe.py --repair  # Re-download Shiller + rebuild")
